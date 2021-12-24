# -*- coding: utf-8 -*-

import re
import ply.lex as lex
import ply.yacc as yacc
import sys
import traceback
import networkx as nx
import openpyxl
import os
import json
import sys
import inspect
from enum import Enum
import yaml

# project modules
from networkml.error import NetworkError, NetworkNotImplementationError
from networkml.generic import GenericValueHolder, Comparator, GenericComponent, GenericValidator, GenericValidatorParam
from networkml.generic import GenericEvaluatee
from networkml.network import ReachabilitySpecification, Numberset, CommandOption, DefaultSorter, NetworkSymbol
from networkml.network import NetworkReturnValue, NetworkNothing
from networkml.validator import GenericEvaluatee, UnaryEvaluatee, BinaryEvaluatee, TrueEvaluatee
import networkml.genericutils as GU
from networkml.generic import debug


def represent_odict(dumper, instance):
    return dumper.represent_mapping('tag:yaml.org,2002:map', instance.items())

yaml.add_representer(dict, represent_odict)

def construct_odict(loader, node):
    return dict(loader.construct_pairs(node))

yaml.add_constructor('tag:yaml.org,2002:map', construct_odict)


class SpecValidator(GenericComponent, GenericValidator):

    def __init__(self, owner=None):
        super().__init__(owner)
        self._validatee = None
        self._validation_policy = GenericValidatorParam.VALIDATE_AS_VARABLE
        self._exception_policy = GenericValidatorParam.ReportExceptionReasons
        self._dic = None
        self._prev_policy = self._validation_policy

    @property
    def validatee(self):
        return self._validatee

    def set_evaluation_target(self, target, policy):
        self._prev_policy = self._validation_policy
        self._validation_policy = policy
        self._dic = target

    def reverse_evaluation_policy(self):
        self._validation_policy = self._prev_policy

    def set_owner(self, owner):
        super().set_owner(owner)
        if self.validatee is not None:
            self.validatee.set_owner(owner)

    def set_validatee(self, validatee):
        self._validatee = validatee
        self._validatee.set_validator(self)

    def set_validation_policy(self, policy: GenericValidatorParam):
        self._validation_policy = policy

    def set_exception_policy(self, policy: GenericValidatorParam):
        self._exception_policy = policy

    def reset_evaluation_policy(self):
        self._validation_policy = GenericValidatorParam.VALIDATE_NOTHING

    def validate(self, evaluatee):
        if isinstance(evaluatee, NetworkSymbol):
            validatee = evaluatee.symbol
        if self._validation_policy == GenericValidatorParam.VALIDATE_NOTHING:
            return evaluatee
        elif self._validation_policy == GenericValidatorParam.VALIDATE_AS_TAG:
            if evaluatee in self._dic.keys():
                return self._dic[evaluatee]
            else:
                return None
        elif self._validation_policy == GenericValidatorParam.VALIDATE_AS_VARABLE:
            var = self.owner.accessor.get(self.owner, evaluatee)
            return var
        else:
            raise NetworkError("Unknown validation policy:{}".format(self._validation_policy))

    def unary_validate(self, value):
        if isinstance(value, NetworkSymbol):
            value = value.symbol
        if self._validation_policy == GenericValidatorParam.VALIDATE_NOTHING:
            return value
        elif self._validation_policy == GenericValidatorParam.VALIDATE_AS_TAG:
            if value in self._dic:
                return self._dic[value]
            else:
                return None
        elif self._validation_policy == GenericValidatorParam.VALIDATE_AS_VARABLE:
            var = self.owner.accessor.get(self, value)
            return var

    def binary_validate(self, l, r):
        if isinstance(l, NetworkSymbol):
            l = l.symbol
        if isinstance(r, NetworkSymbol):
            r = r.symbol
        if self._validation_policy == GenericValidatorParam.VALIDATE_NOTHING:
            return l, r
        elif self._validation_policy == GenericValidatorParam.VALIDATE_AS_TAG:
            if l in self._dic.keys():
                return self._dic[l], r
            return l, None
        elif self._validation_policy == GenericValidatorParam.VALIDATE_AS_VARABLE:
            var = self.owner.accessor.get(self, r)
            r = var
            return l, r

    def validate_left(self, l):
        pass

    def validate_right(self, r):
        pass


class SpecificationGraphParam(Enum):

    CONTEXT = "context"
    VARIABLES = "variables"
    FUNCTIONS = "functions"
    FUNCTION = "function"
    ARITY = "arity"
    CAPTION = "caption"
    ARGTYPES = "argtypes"
    DOC = "doc"


class SpecificationGraph(nx.MultiDiGraph):

    Arity = SpecificationGraphParam.ARITY
    Function = SpecificationGraphParam.FUNCTION
    Caption = SpecificationGraphParam.CAPTION
    ArgTypes = SpecificationGraphParam.ARGTYPES
    Doc = SpecificationGraphParam.DOC
    CONSTRUCT_SPEC_NETWORK = "construct"
    CHECK_REACHABILITY = "reachability"

    _spec_doc = None

    @property
    def spec_doc(self):
        return self._spec_doc

    def __init__(self, N: nx.MultiDiGraph, config_file, filename=None, traceback=False, enable_stack=True):
        super().__init__()
        self._N = None
        self._config_file = config_file
        self._filename = None
        self._traceback = traceback
        if enable_stack is None:
            self._enable_stack = False
        else:
            self._enable_stack = enable_stack
        self.init(N, filename)
        self._validator = SpecValidator()
        if self._spec_doc is None:
            self.construct_spec_doc()

    def init(self, N: nx.MultiDiGraph, filename=None):
        self._N = N
        self._filename = filename

    @property
    def validator(self):
        return self._validator

    def set_validator(self, validator):
        self._validator = validator

    # FIXME
    def construct_spec_doc(self):
        descfile = self._config_file
        try:
            xml = GU.read_xml(descfile)
            root = xml.getroot()
            self._spec_doc = root
        except Exception as ex:
            debug(ex)
            debug("{} not found. We cannot provide document of our class.".format(descfile))

    def load_settings(self):
        self.construct_spec_doc()

    def reload_settings(self):
        self.load_settings()

    def find_method_doc(self, method_sig, doc_kind):  # doc_kind in {"doc", "args-requirement"}
        for m in self._spec_doc:
            # m.text, m.attrib, m.tag
            if m.tag == "method" and m.attrib["name"] == method_sig:
                for n in m:
                    if n.tag == doc_kind:
                        return n.text
        return None

    def arrange_args(self, caller, sig, args):
        arrange_doc = self.find_method_doc(sig, "args-arrangement")
        if arrange_doc is None:
            return args
        doc = arrange_doc.split("\n")
        tags_list = []
        for i, d in enumerate(doc):
            d = d.replace(" ", "")
            d = d.replace("\t", "")
            tags = d.split(",")
            if tags[0] != "":
                tags_list.append(tags)
        # At first, generates None value options
        new_args = [CommandOption("--{}".format(ts[0]), None, has_assignee=False) for ts in tags_list]
        for a in args:
            puts = False
            for i, tags in enumerate(tags_list):
                if a.name in tags:
                    new_args[i] = a
                    puts = True
                    break
            if not puts:
                debug("{} was ignored.".format(a))
        return new_args

    def build_args_checker(self, caller, args, method_sig):
        program = []
        i = -1
        skip_pat = r"^\s*(#.*|//.*|)$"
        ava_pat = r"^\s*refer\s*=\s*(?P<actual>.*)\s*$"
        if_pat = r"^\s*if\s+(?P<condition>.+)(\s*|\s*:\s*)$"
        endif_pat = r"^\s*fi\s*$"
        # eq_pat = r"^\s*(?P<eq>.*)\s*\->\s*(?P<result>(True|False))\s*(|,(?P<message>.*))\s*$"
        explicit_ret = r"^\s*return\s+(?P<return>(True|False))(|,(?P<message>.*))\s*$"
        ret_pat = r"^\s*(?P<return>(True|False))(|,(?P<message>.*))\s*$"
        if_depth = []
        step = None
        try:
            doc = self.find_method_doc(method_sig, "args-requirement")
            lines = doc.split("\n")
            for i, e in enumerate(lines):
                step = {}
                program.append(step)
                step['sentence'] = e
                m = re.match(skip_pat, e)
                if m is not None:
                    step['statement'] = "ignore"
                    step['condition'] = None
                    continue
                m = re.match(ava_pat, e)
                if m is not None:
                    step['statement'] = "refer"
                    d = m.groupdict()
                    p = d['actual']
                    step['condition'] = True
                    step['jump'] = p
                    continue
                m = re.match(if_pat, e)
                if m is not None:
                    step['statement'] = "if"
                    d = m.groupdict()
                    p = d['condition']
                    p = "lambda caller, args: {}".format(p)
                    p = eval(p)
                    step['condition'] = p
                    step['jump'] = i+1
                    step['else'] = None
                    if_depth.append(step)
                    continue

                m = re.match(endif_pat, e)
                if m is not None:
                    step['statement'] = "endif"
                    if len(if_depth) > 0:
                        if_step = if_depth[len(if_depth)-1]
                        if_step['else'] = i+1
                        if_depth.remove(if_step)
                    continue

                m = re.match(ret_pat, e)
                if m is not None:
                    step['statement'] = "return"
                    d = m.groupdict()
                    p = d['return']
                    q = d['message']
                    result = eval(p)
                    message = q
                    step['condition'] = True
                    step['return'] = result
                    step['message'] = message
                    if len(if_depth) > 0:
                        if_step = if_depth[len(if_depth)-1]
                        if_step['else'] = i+1
                        if_depth.remove(if_step)
                        # print(if_step)
                    continue
                m = re.match(explicit_ret, e)
                if m is not None:
                    step['statement'] = "return"
                    d = m.groupdict()
                    p = d['return']
                    q = d['message']
                    result = eval(p)
                    message = q
                    step['condition'] = True
                    step['return'] = result
                    step['message'] = message
                    if len(if_depth) > 0:
                        if_step = if_depth[len(if_depth)-1]
                        if_step['else'] = i+1
                        if_depth.remove(if_step)
                        # print(if_step)
                    continue
                debug("syntax error at line {} ignored.:{}".format(i, e))
            return program
        except Exception as ex:
            debug("{}: error. {}".format(i, step))
        finally:
            pass

    def check_args_requirement(self, caller, method_sig, args):
        # if --help, print help.
        if len(args) > 0 and isinstance(args[0], CommandOption) and args[0].name == "help":
            self.help_check(method_sig, caller, args)
            rtn = NetworkReturnValue(args, True, "", cancel=True)
            return rtn

        # solve symbols
        new_args = []
        for a in args:
            if isinstance(a, CommandOption):
                if a.has_assignee:
                    value = a.value
                    if isinstance(value, NetworkSymbol):
                        value = caller.accessor.get(caller, value.symbol)
                    opt = CommandOption("--{}".format(a.name), value, has_assignee=True, symbolic_assignee=False)
                    a = opt
            new_args.append(a)
        args = new_args

        # arrange arguments
        args = self.arrange_args(caller, method_sig, args)

        checker = self.build_args_checker(caller, args, method_sig)
        i = 0
        while True:
            step: dict = checker[i]
            # print("{}: {}".format(i, step))
            if step['statement'] == "ignore":
                i = i+1
                continue
            elif step['statement'] == "endif":
                i = i+1
                continue
            elif step['statement'] == "return":
                break
            elif step['statement'] == "if":
                cond = step['condition'](caller, args)
                if cond:
                    i = step['jump']
                else:
                    i = step['else']
                continue
            elif step['statement'] == 'refer':
                checker = self.build_args_checker(caller, args, step['jump'])
                i = 0
                continue

        ret = step['return']
        msg = step['message']
        rtn = NetworkReturnValue(args, ret, msg, cancel=not ret)
        return rtn

    @property
    def enable_stack(self):
        return self._enable_stack

    @property
    def traceback(self):
        return self._traceback

    def help_check(self, sig, caller, args):
        if len(args) > 0:
            help = args[0]
            if not isinstance(help, CommandOption):
                return None
            if help.name != "help":
                return None
            doc = self.find_method_doc(sig, "doc")
            eqs = self.find_method_doc(sig, "args-requirement")
            if not args[0].has_assignee:
                print_doc = True
                print_req = True
            elif args[0].value == "doc":
                print_doc = True
                print_req = False
            elif args[0].value == "args":
                print_doc = False
                print_req = True
            else:
                debug("--help=(doc|requirement)")
                print_doc = True
                print_req = True
            caller.print_buf(sig)
            if print_doc:
                caller.print_buf("doc:")
                caller.print_buf(doc)
            if print_req:
                caller.print_buf(eqs)
            rtn = NetworkReturnValue(args, True, "", cancel=True)
            return rtn
        else:
            return None

    def collect_nodes_with_candidates(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("collect_nodes_with_candidates", caller, args)
                if ret is not None:
                    return ret
            default_spec = TrueEvaluatee(None)
            if args[0].has_assignee:
                node_spec = args[0].value
            else:
                node_spec = default_spec
            if args[1].has_assinee:
                node_set = args[1].value
            else:
                node_set = ()
            with_data = False
            if len(args) > 2:
                if not args[2].has_assignee:
                    with_data = args[2].value
            N = []
            for n in self.N.nodes(caller, data=True):
                if node_spec.evaluate(caller):
                    if n in node_set:
                        if with_data:
                            N.append(n)
                        else:
                            N.append(n[0])
            return N
        except Exception as ex:
            raise NetworkError("collect_nodes_with_candidates failed:{}".format(args), ex)

    def collect_edges_with_candidates(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("collect_edges_with_candidates", caller, args)
                if ret is not None:
                    return ret
            none_spec = TrueEvaluatee(None)
            src_spec = args[0].value
            if not args[0].has_assignee:
                src_spec = none_spec
            src_nodes = args[0].value
            if not args[1].has_assignee:
                src_nodes = ()
            dst_spec = args[2].value
            if not args[2].has_assignee:
                dst_spec = none_spec
            dst_nodes = args[3].value
            if not args[3].has_assignee:
                dst_nodes = ()
            edge_spec = args[4].value
            if not args[4].has_assignee:
                edge_spec = none_spec
            edges = args[5].value
            if not args[5].has_assignee:
                edges = ()
            if len(args) > 6:
                if args[6].has_assignee:
                    with_data = args[6].value
                else:
                    with_data = False
            else:
                with_data = False
            if len(dst_nodes) == 0:
                dst_nodes = [_ for _ in self.N.nodes]
            if len(src_nodes) == 0:
                src_nodes = [_ for _ in self.N.nodes]
            if len(edges) == 0:
                edges = [_ for _ in self.N.edges(keys=True, data=True)]
            if caller.validator is None:
                validator = self.validator
            else:
                validator = caller.validator
            E = []
            for u in src_nodes:  # for each u in U
                if isinstance(validator, SpecValidator):
                    validator.set_evaluation_target(self.N.nodes[u], GenericValidatorParam.VALIDATE_AS_TAG)
                result = src_spec.evaluate(caller)
                if isinstance(validator, SpecValidator):
                    validator.reverse_evaluation_policy()
                if not result:
                    continue
                for v in dst_nodes:  # for each v in V
                    if isinstance(validator, SpecValidator):
                        validator.set_evaluation_target(self.N.nodes[v], GenericValidatorParam.VALIDATE_AS_TAG)
                    result = dst_spec.evaluate(caller)
                    if isinstance(validator, SpecValidator):
                        validator.reverse_evaluation_policy()
                    if not result:
                        continue
                    for e in edges:  # for each e in E
                        if u != e[0] or v != e[1]:
                            continue
                        for k in self.N[u][v].keys():
                            # e = self.N[u][v][k]
                            if isinstance(validator, SpecValidator):
                                validator.set_evaluation_target(self.N[u][v][k], GenericValidatorParam.VALIDATE_AS_TAG)
                            result = dst_spec.evaluate(caller)
                            if isinstance(validator, SpecValidator):
                                validator.reverse_evaluation_policy()
                            if not result:
                                continue
                            if with_data:
                                E.append((u, v, k, self.N[u][v][k]))
                            else:
                                E.append((u, v, k))
            return E
        except Exception as ex:
            raise NetworkError("collect_edges_with_candidates failed:{}".format(args), ex)

    def node_attr_names(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("node_attr_names", caller, args)
                if ret is not None:
                    return ret
            node_set = ()
            if len(args) > 0:
                node_set = args[0].value
                if not args[0].has_assignee:
                    node_set = ()
            if len(node_set) == 0:
                node_set = [_ for _ in self.N.nodes]
            attr_names = []
            for n in node_set:
                for k in self.N.nodes[n].keys():
                    if k not in attr_names:
                        attr_names.append(k)
            return attr_names
        except Exception as ex:
            raise NetworkError("node_attr_names failed:{}".format(args), ex)

    def edge_attr_names(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("edge_attr_names", caller, args)
                if ret is not None:
                    return ret
            edge_set = ()
            if len(args) > 0:
                edge_set = args[0].value
                if not args[0].has_assignee:
                    edge_set = ()
            if len(edge_set) == 0:
                edge_set = [_ for _ in self.N.edges]
            attr_names = []
            for e in edge_set:
                for k in self.N[e[0]][e[1]].keys():
                    attr = self.N[e[0]][e[1]][k]
                    for m in attr.keys():
                        if m not in attr_names:
                            attr_names.append(m)
            return attr_names
        except Exception as ex:
            raise NetworkError("collect_edge_attr_names failed:{}".format(args), ex)

    def node_attr_values(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("node_attr_values", caller, args)
                if ret is not None:
                    return ret
            if args[0].has_assignee:
                node_specs = args[1].value
            else:
                node_specs = ()
            if args[1].has_assignee:
                node_set = args[1].value
            else:
                node_set = [_ for _ in self.N.nodes]
            attr_values = []
            for n in node_set:
                for spec in node_specs:
                    if spec in self.N.nodes[n].keys():
                        if self.N.nodes[n][spec.l] not in attr_values:
                            attr_values.append(self.N.nodes[n][spec.l])
            return attr_values
        except Exception as ex:
            raise NetworkError("node_attr_values failed:{}".format(args), ex)

    def edge_attr_values(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("edge_attr_names", caller, args)
                if ret is not None:
                    return ret
            attr = args[0].value
            edge_spec = args[1].value
            edge_set = None
            if len(args) > 2:
                edge_spec = args[2]
            if len(args) > 3:
                edge_set = args[3]
            if len(edge_set) == 0:
                edge_set = [_ for _ in self.N.edges]
            attr_values = []
            for e in edge_set:
                for k in self.N[e[0]][e[1]].keys():
                    if k in edge_spec:
                        if self.N[e[0]][e[1]][k][attr] not in attr_values:
                            attr_values.append(self.N[e[0]][e[1]][k][attr])
            return attr_values
        except Exception as ex:
            raise NetworkError("edge_attr_values failed:{}".format(args), ex)

    def resolve_symbol(self, caller, sym):
        if isinstance(sym, NetworkSymbol):
            return sym.value
        return sym

    def nodes_product(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("nodes_product", caller, args)
                if ret is not None:
                    return ret
            U = args[0].value
            U = self.resolve_symbol(caller, U)
            V = args[1].value
            V = self.resolve_symbol(caller, V)
            node_product = []
            for u in U:
                for v in V:
                    node_product.append((u, v))
            return node_product
        except Exception as ex:
            raise NetworkError("nodes_product failed:{}".format(args), ex)

    def edges_to_nodes_product(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("edges_to_nodes_product", caller, args)
                if ret is not None:
                    return ret
            node_product = []
            edges = args[0].value
            for e in edges:
                if (e[0], e[1]) not in node_product:
                    node_product.append((e[0], e[1]))
            return node_product
        except Exception as ex:
            raise NetworkError("edges_to_nodes_product failed:{}".format(args), ex)

    def subtract_nodes_product(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("subtract_nodes_product", caller, args)
                if ret is not None:
                    return ret
            U = args[0].value
            V = args[1].value
            N = []
            for u in U:
                if u not in V:
                    N.append(u)
            return N
        except Exception as ex:
            raise NetworkError("subtract_nodes_product failed:{}".format(args), ex)

    def nodes_product_to_edges(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("nodes_product_to_edges", caller, args)
                if ret is not None:
                    return ret
            P = args[0].value
            attrs = args[1].value
            if not args[1].has_assignee:
                attrs = ()
            if args[2].has_assignee:
                overwrite = args[2].value
            else:
                overwrite = False
            E = []
            for e in P:
                if not (e[0] in self.N.nodes or e[1] in self.N[e[0]].keys()):
                    if not overwrite:
                        continue
                else:
                    self.N.add_edge(e[0], e[1])
                    E.append((e[0], e[1], attrs))
                for k in self.N[e[0]][e[1]].keys():
                    e = self.N[e[0]][e[1]][k]
                    for a in attrs:
                        e[a.l] = a.r
                # o1 = CommandOption("-u", [e[0]], has_assignee=True)
                # o2 = CommandOption("-v", [e[1]], has_assignee=True)
                # o3 = CommandOption("-a", attrs, has_assignee=True)
                # o4 = CommandOption("-ow", overwrite, has_assignee=True)
                # self.newedge(caller, (o1, o2, o3, o4))
                # E.append((e[0], e[1], attrs))
            return E
        except Exception as ex:
            raise NetworkError("nodes_product_to_edges failed:{}".format(args), ex)

    def project(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("project", caller, args)
                if ret is not None:
                    return ret
            src_set = args[0].value
            numbers = args[1].value
            dst_set = []
            for s in src_set:
                d = []
                for i in numbers:
                    d.append(s[i])
                if d not in dst_set:
                    dst_set.append(d)
            return dst_set
        except Exception as ex:
            raise NetworkError("project failed:{}".format(args), ex)

    @property
    def N(self) -> nx.MultiDiGraph:
        return self._N

    @property
    def filename(self) -> str:
        return self._filename

    def analyze_reach_segments(self, caller, spec: ReachabilitySpecification, S, D):
        result = []
        edge_dict = spec.edge_dict
        for seg_idx, seg in enumerate(sorted(edge_dict.keys())):
            seg_idx = seg_idx+1
            if not S.contains(seg_idx):
                continue
            depths = edge_dict[seg][ReachabilitySpecification.DEPTHS]
            for dep in depths:
                if not D.contains(dep):
                    continue
                Q = edge_dict[seg][ReachabilitySpecification.QUANTIFIER]
                if Q.contains(dep):
                    if depths[dep][ReachabilitySpecification.DST_REACHABLE]:
                        r, m = self.analyze_reachable(caller, spec, seg, dep)
                        result.append((r, {"seg": seg, "depth": dep, "message": m, "reachable": True}))
                    else:
                        result.append((False, {"seg": seg, "depth": dep, "message": "", "reachable": False}))
                else:
                    if depths[dep][ReachabilitySpecification.DST_REACHABLE]:
                        result.append((True, {"seg": seg, "depth": dep, "message": "not in specification point.",
                                              "reachable": True}))
                    else:
                        result.append((False, {"seg": seg, "depth": dep, "message": "not in specification point.",
                                               "reachable": False}))
        return result

    def analyze_reachable(self, caller, spec: ReachabilitySpecification, seg, dep):
        try:
            if seg not in spec.edge_dict.keys():
                return False, "Invalid edge key:{}".format(seg)
            Q = spec.edge_dict[seg][ReachabilitySpecification.QUANTIFIER]
            depths = spec.edge_dict[seg][ReachabilitySpecification.DEPTHS]
            if dep not in depths.keys():
                return False, "Invalid depth {} in segment {}.".format(dep, seg)
            if not Q.contains(dep):
                return False, "Depth {} is not in available quantifier of segment {}".format(dep, seg)
            if depths[dep][ReachabilitySpecification.DST_REACHABLE]:
                return True, "Reachable to segment {}, depth {}".format(seg, dep)
            else:
                return False, "Depth {} is not in available quantifier of segment {}".format(dep, seg)

        except Exception as ex:
            raise NetworkError("reachable failed. spec:{}, seg:{}, depth:{}".format(spec, seg, dep), ex)

    def analyze_reach_loopback(self, caller, spec: ReachabilitySpecification, _from, _to):
        raise NetworkNotImplementationError("{} not implementd.".format("analyze_reach_loopback"))
        pass

    def check_reachability(self, caller, spec: ReachabilitySpecification, seg, depth, loopback, _from, _to):
        if seg is not None:
            return self.analyze_reach_segments(caller, spec, seg, depth)
        elif loopback is not None:
            return self.analyze_reach_loopback(caller, spec, _from, _to)
        else:
            debug("Invalid arguments")
            return None
        #     segment = len(keys) - 1
        #     S = Numberset(1)
        #     if len(options) > 0:
        #         if options[0].value == "last":
        #             segment = len(keys) - 1
        #             S = Numberset(segment)
        #         elif type(options[0].value) is int:
        #             segment = options[0].value
        #             if segment < 0 or len(keys)-1 < segment:
        #                 return [(None, "invalid segment:{}".format(options))]
        #             S = Numberset(segment)
        #         elif isinstance(options[0].value, Numberset):
        #             S = options[0].value
        #         else:
        #             return [(None, "invalid segment:{}".format(options))]
        #
        #     branch = None
        #     B = Numberset(1)
        #     if len(options) > 1:
        #         if options[1].value == "last":
        #             branch = "last"
        #         elif type(options[1].value) is int:
        #             B = Numberset(options[1].value)
        #         elif isinstance(options[1].value, Numberset):
        #             B = options[1].value
        #         else:
        #             return [(None, None, None, "invalid branch:{}".format(options))]
        #
        #     seg = S.minimum
        #     while seg is not None and S.contains(seg) and seg < len(keys):
        #         quantifier = spec.edge_dict[keys[seg]][ReachabilitySpecification.QUANTIFIER]
        #         depths = spec.edge_dict[keys[seg]][ReachabilitySpecification.DEPTHS]
        #         if branch == "last":
        #             branch_keys = sorted(depths.keys())
        #             if len(branch_keys) != 0:
        #                 B = Numberset(branch_keys[0])
        #             else:
        #                 pass
        #         branch = B.minimum
        #         while branch is not None and B.contains(branch):
        #             if quantifier.contains(branch) and depths[branch][ReachabilitySpecification.DST_REACHABLE]:
        #                 r, m = self.reachable(caller, spec, seg, branch)
        #                 result.append((r, seg, branch, m))
        #             else:
        #                 result.append((False, seg, branch, ""))
        #             branch = B.succ(branch)
        #         seg = S.succ(seg)
        #     return result
        # except Exception as ex:
        #     raise NetworkError("check_reachability failed:{}".format(spec, options), ex)

    def reachable(self, caller, spec: ReachabilitySpecification, seg, dep):
        try:
            keys = sorted(spec.edge_dict.keys())
            if seg < 0 or len(keys)-1 < seg:
                return False, "invalid segment:{}".format(seg)
            if seg >= len(keys):
                return False, "segment {} is out of bounds of in key.".format(seg)
            Q = spec.edge_dict[keys[seg]][ReachabilitySpecification.QUANTIFIER]
            depths = spec.edge_dict[keys[seg]][ReachabilitySpecification.DEPTHS]
            if dep in depths.keys():
                if Q.contains(dep):
                    if depths[dep][ReachabilitySpecification.DST_REACHABLE]:
                        return True, "Reachable to segment {}, depth {}".format(seg, dep)
                    else:
                        return False, "Unreachable to segment {}, depth {}".format(seg, dep)
                else:
                    return False, "Depth {} is not in available quantifier of segment {}".format(dep, seg)
            else:
                return False, "Depth {} is unavailable in segment {}".format(dep, seg)
        except Exception as ex:
            raise NetworkError("reachable failed:{}".format(spec, seg, dep), ex)

    def construct_spec_network(self, caller, spec: ReachabilitySpecification):
        try:
            edge_connection = spec.serialize()
            dic = {}
            spec.edge_dictionary(dic, edge_connection, [], "TOP")
            spec._edge_dict = dic
            edge_dict = spec.edge_dict
            # construct forward network
            for seg in sorted(edge_dict.keys()):
                self.collect_edge_dst_nodes(caller, spec, seg)
        except Exception as ex:
            raise NetworkError("construct_spec_network failed:{}".format(spec), ex)

    def collect_edge_eventual_dst_nodes(self, caller, spec: ReachabilitySpecification, seg):
        try:
            eventual_dst_nodes = []
            edge_dict = spec.edge_dict
            quantifier = edge_dict[seg][ReachabilitySpecification.QUANTIFIER]
            depths = edge_dict[seg][ReachabilitySpecification.DEPTHS]
            for i in depths.keys():
                if quantifier.contains(i):
                    dst_nodes = depths[i][ReachabilitySpecification.DST_NODES]
                    for n in dst_nodes:
                        if n not in eventual_dst_nodes:
                            eventual_dst_nodes.append(n)
            return eventual_dst_nodes
        except Exception as ex:
            raise NetworkError("collect_edge_eventual_dst_nodes failed:{}".format(spec, seg), ex)

    def collect_edge_dst_nodes(self, caller, spec: ReachabilitySpecification, seg):
        try:
            edge_dict = spec.edge_dict
            keys = sorted(edge_dict.keys())
            seg_idx = keys.index(seg)
            src_spec = edge_dict[seg][ReachabilitySpecification.SRC_SPEC]
            edge_spec = edge_dict[seg][ReachabilitySpecification.EDGE_SPEC]
            dst_spec = edge_dict[seg][ReachabilitySpecification.DST_SPEC]
            quantifier = edge_dict[seg][ReachabilitySpecification.QUANTIFIER]
            if seg_idx == 0:  # toplevel
                opt_spec = CommandOption("-spec", [src_spec], has_assignee=True)
                opt_cand = CommandOption("-candidates", has_assignee=False)
                opt_data = CommandOption("-with_data", has_assignee=False)
                src_nodes = self.collect_nodes(caller, (opt_spec, opt_cand, opt_data))
            else:
                prev_seg = keys[seg_idx-1]
                src_nodes = self.collect_edge_eventual_dst_nodes(caller, spec, prev_seg)
            edge_dict[seg][ReachabilitySpecification.DEPTHS] = {}
            for i in range(1, quantifier.maximum+1):
                depths = edge_dict[seg][ReachabilitySpecification.DEPTHS]
                depths[i] = {}
                depths[i][ReachabilitySpecification.SRC_NODES] = src_nodes
                depths[i][ReachabilitySpecification.LOOPBACK_NODES] = {}
                depths[i][ReachabilitySpecification.SRC_REACHABLE] = len(src_nodes) != 0
                if not depths[i][ReachabilitySpecification.SRC_REACHABLE]:
                    depths[i][ReachabilitySpecification.DST_NODES] = []
                    depths[i][ReachabilitySpecification.DST_REACHABLE] = False
                    break
                opt_src_nodes = CommandOption("-src_nodes", src_nodes, has_assignee=True)
                opt_edge_spec = CommandOption("-edge_spec", [edge_spec], has_assignee=True)
                opt_dst_spec = CommandOption("-dst_spec", [dst_spec], has_assignee=True)
                opt_data = CommandOption("-with_data", has_assignee=False)
                dst_nodes = self.collect_dst_nodes(caller, (opt_src_nodes, opt_edge_spec, opt_dst_spec, opt_data))
                depths[i][ReachabilitySpecification.DST_NODES] = dst_nodes
                depths[i][ReachabilitySpecification.DST_REACHABLE] = len(dst_nodes) != 0
                if not depths[i][ReachabilitySpecification.DST_REACHABLE]:  # stop here, forward disabled
                    break
                src_nodes = dst_nodes
                # construct recursion network
                for dest_seg_idx in range(seg_idx+1):
                    m = sorted(edge_dict.keys())[dest_seg_idx]
                    dest_depths = edge_dict[m][ReachabilitySpecification.DEPTHS]
                    for j in sorted(dest_depths.keys()):
                        back_src_nodes = dest_depths[j][ReachabilitySpecification.SRC_NODES]
                        loopback_nodes = []
                        for n in dst_nodes:
                            if n in back_src_nodes:
                                loopback_nodes.append(n)
                        depths[i][ReachabilitySpecification.LOOPBACK_NODES][(dest_seg_idx, j)] = loopback_nodes
        except Exception as ex:
            raise NetworkError("collect_edge_dst_nodes:({},{})".format(spec, seg), ex)

    def select_edges(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("select_edges", caller, args)
            if ret is not None:
                return ret
        return self.collect_edges(caller, args)

    def collect_edges(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("collect_edges", caller, args)
            if ret is not None:
                return ret
        if args[0].has_assignee:
            edge_specs = args[0].value
        else:
            edge_specs = ()
        if args[1].has_assignee:
            D = args[1].value
        else:
            D = None
        if args[2].has_assignee:
            with_data = args[2].value
        else:
            with_data = False
        if D is None:
            D = [_ for _ in self.N.edges(keys=True, data=with_data)]
        E = []
        # validator setup
        if caller is None:
            validator = self.validator
        else:
            validator = caller.validator
        for e in D:
            done = True
            for a in edge_specs:
                if isinstance(a, GenericEvaluatee):
                    validator.set_evaluation_target(e[3], GenericValidatorParam.VALIDATE_AS_TAG)
                    result = a.evaluate(caller)
                    validator.reverse_evaluation_policy()
                else:
                    result = a.value
                if not result:
                    done = False
                    break
            if done:
                if with_data:
                    E.append(e)
                else:
                    E.append((e[0], e[1], e[2]))
        return E

    def collect_args(self, caller, args, type_list=()):
        holder = [[] for _ in type_list]
        for a in args:
            for i, t in enumerate(type_list):
                if isinstance(a, t):
                    holder[i].append(a)
        return holder

    def select_nodes(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("select_nodes", caller, args)
            if ret is not None:
                return ret
        if args[0].has_assignee:
            node_specs = args[0].value
        else:
            node_specs = ()
        if args[1].has_assignee:
            N = args[1].value
        else:
            N = None
        if N is None:
            N = [_ for _ in self.N.nodes]
        if args[2].has_assignee:
            with_data = args[2].value
        else:
            with_data = False
        V = []
        # validator setup
        if caller is None:
            validator = self.validator
        else:
            validator = caller.validator
        for n in N:
            done = True
            for a in node_specs:
                if isinstance(a, GenericEvaluatee):
                    validator.set_evaluation_target(self.N.nodes[n], GenericValidatorParam.VALIDATE_AS_TAG)
                    result = a.evaluate(caller)
                    validator.reverse_evaluation_policy()
                else:
                    result = a.value
                if not result:
                    done = False
                    break
            if done:
                if with_data:
                    V.append((n, self.N.nodes[n]))
                else:
                    V.append(n)
        return V

    def collect_nodes(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("collect_nodes", caller, args)
            if ret is not None:
                return ret
        return self.select_nodes(caller, args)

    def collect_dst_nodes(self, caller, args):
        try:
            ret = self.help_check("collect_dst_nodes", caller, args)
            if ret is not None:
                return ret
            if args[0].has_assignee:
                src_nodes = args[0].value
            else:
                src_nodes = ()
            if args[1].has_assignee:
                edge_specs = args[1].value
            else:
                edge_specs = ()
            if args[2].has_assignee:
                dst_specs = args[2].value
            else:
                dst_specs = ()
            with_data = False
            if len(args) > 3:
                if args[3].has_assignee:
                    with_data = args[3].value
            D = []
            if caller.validator is None:
                validator = self.validator
            else:
                validator = caller.validator
            for u in src_nodes:
                for v in self.N[u].keys():
                    done = True
                    for a in dst_specs:
                        if isinstance(validator, SpecValidator):
                            validator.set_evaluation_target(self.N.nodes[v], GenericValidatorParam.VALIDATE_AS_TAG)
                        result = a.evaluate(caller)
                        if isinstance(validator, SpecValidator):
                            validator.reverse_evaluation_policy()
                        if not result:
                            done = False
                            break
                    if not done:
                        continue
                    for k in self.N[u][v].keys():
                        done = True
                        for a in edge_specs:
                            if isinstance(validator, SpecValidator):
                                validator.set_evaluation_target(self.N[u][v][k], GenericValidatorParam.VALIDATE_AS_TAG)
                            result = a.evaluate(caller)
                            if isinstance(validator, SpecValidator):
                                validator.reverse_evaluation_policy()
                            if not result:
                                done = False
                                break
                        if not done:
                            continue
                        if v not in D:
                            if with_data:
                                D.append((v, self.N.nodes[v]))
                            else:
                                D.append(v)
                            break
            return D
        except Exception as ex:
            raise NetworkError("collect_dst_nodes failed:{}".format(args), ex)

    def cardinality(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("cardinality", caller, args)
                if ret is not None:
                    return ret
            if not args[0].has_assignee:
                # ignore
                return
            container = args[0].value
            if isinstance(container, NetworkSymbol):
                container = container.value
            return len(container)
            # raise NetworkError("argument error:{}".format(a))
        except Exception as ex:
            raise NetworkError("cardinality failed.")

    def minus(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("minus", caller, args)
            if ret is not None:
                return ret
        x = args[0]
        y = args[1]
        if type(x) is list and type(y) is list:
            z = x.copy()
            for a in y:
                if a in z:
                    z.remove(a)
            return z
        elif type(x) is list and type(y) is not list:
            z = x.copy()
            if y in x:
                z.remove(y)
            return z
        if (type(x) is int or type(x) is float) and (type(y) is int or type(y) is float):
            z = x - y
            return z
        return None

    def new_node_id(self, caller):
        n = 0
        for n in range(1, len(self.N.nodes)+1):
            if n not in self.N.nodes:
                return n
        return n+1

    def new_node_with_constraint_args(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("new_node_with_constraint_args", caller, args)
                if ret is not None:
                    return ret
            if args[0].has_assignee:
                specs = args[0].value
            else:
                specs = ()
            debug("specs:", specs)
            if args[1].has_assignee:
                candidates = args[1].value
            else:
                candidates = ()
            if args[2].has_assignee:
                overwrite = args[2].value
            else:
                overwrite = False
            if len(candidates) == 0:
                candidates = (self.new_node_id(caller))
            debug("candidatates:", candidates)
            for n in candidates:
                if n in self.N.nodes and not overwrite:
                    continue
                self.N.add_node(n)
                for a in specs:
                    if isinstance(a, BinaryEvaluatee):
                        self.N.nodes[n][a.l] = a.r
            return candidates
        except Exception as ex:
            raise NetworkError("deledges failed:{}".format(args), ex)

    def newnode(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("newnode", caller, args)
            if ret is not None:
                return ret
        if args[0].has_assignee:
            S = args[0].value
        else:
            S = ()
        if args[1].has_assignee:
            C = args[1].value
        else:
            C = [self.new_node_id(caller)]

        if args[2].has_assignee:
            ow = args[2].value
        else:
            ow = False
        N = []
        for c in C:
            if c not in self.N.nodes:
                exists = False
                self.N.add_node(c)
                N.append(c)
            else:
                exists = False
            if ow or not exists:
                for s in S:
                    self.N.nodes[c][s.l] = s.r
        return N

    def newnodes(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("newnodes", caller, args)
            if ret is not None:
                return ret
        return self.newnode(caller, args)

    def edge_keys(self, caller, u, v):
        if v in self.N[u].keys():
            return [_ for _ in self.N[u][v].keys()]
        else:
            return []

    def newedge(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("newedge", caller, args)
                if ret is not None:
                    return ret
            U = args[0].value
            V = args[1].value
            if args[2].has_assignee:
                A = args[2].value
            else:
                A = ()
            if args[3].has_assignee:
                overwrite = args[3].value
            else:
                overwrite = False
            if U is None or V is None:
                debug("Insufficient node specified.")
                return
            E = []
            for u in U:
                if u not in self.N.nodes:
                    continue
                for v in V:
                    if v not in self.N.nodes:
                        continue
                    if v in self.N[u].keys():  # edge exists
                        EK = self.N[u][v].keys()
                        if overwrite:
                            for ek in EK:
                                for spec in A:
                                    spec: BinaryEvaluatee = spec
                                    self.N[u][v][ek][spec.l] = spec.r
                    else:
                        self.N.add_edge(u, v)
                        key = 1
                        for key in self.N[u][v].keys():
                            break
                        for spec in A:
                            spec: BinaryEvaluatee = spec
                            self.N[u][v][key][spec.l] = spec.r
                        new_edge = (u, v, key, A)
                        E.append(new_edge)
            return E
        except Exception as ex:
            raise NetworkError("newedge failed:{}".format(args), ex)

    def newedges(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("newedges", caller, args)
            if ret is not None:
                return ret
        return self.nodes_product_to_edges(caller, args)

    def delnodes(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("delnodes", caller, args)
                if ret is not None:
                    return ret
            if args[0].has_assignee:
                spec_opt = args[0]
            else:
                spec_opt = CommandOption("-spec", TrueEvaluatee(None))
            node_can = args[1]
            if node_can.has_assignee:
                candidates = node_can.value
            else:
                candidates = None
            if candidates is None:
                candidates = [_ for _ in self.N.nodes]
                node_can = CommandOption("-candidates", candidates, has_assignee=True)
            data_opt = CommandOption("-data", False, has_assignee=False)
            if candidates is None:
                candidates = self.collect_nodes(caller, (spec_opt, node_can, data_opt))
            for n in candidates:
                self.N.remove_node(n)
            return candidates
        except Exception as ex:
            raise NetworkError("delnodes failed:{}".format(args), ex)

    def deledges(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("deledges", caller, args)
                if ret is not None:
                    return ret
            if args[0].has_assignee:
                spec = args[0]
            else:
                spec = CommandOption("-edge_spec", TrueEvaluatee(None))
            can = args[1]
            if can.has_assignee:
                candidates = can.value
            else:
                candidates = None
            if candidates is None:
                candidates = [_ for _ in self.N.edges]
                can = CommandOption("-edges", candidates)
            data_opt = CommandOption("--with_data", False, has_assignee=True)
            candidates = self.collect_edges(caller, (spec, can, data_opt))
            for e in candidates:
                self.N.remove_edge(e[0], e[1], e[2])
            return candidates
        except Exception as ex:
            raise NetworkError("deledges failed:{}".format(args), ex)

    def trivials(self, caller, *args):
        try:
            if len(args) > 0:
                ret = self.help_check("tivials", caller, args)
                if ret is not None:
                    return ret
            triv = []
            for n in self.N.nodes:
                no_edge = True
                for e in self.N.edges:
                    if n == e[0] or n == e[1]:
                        no_edge = False
                        break
                if no_edge:
                    triv.append(n)
            return triv
        except Exception as ex:
            raise NetworkError("trivials failed", ex)

    def len(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("len", caller, args)
            if ret is not None:
                return ret
        return len(args[0])

    def print(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("print", caller, args)
            if ret is not None:
                return ret
        for arg in args:
            debug(arg)
            caller.print_buf.append(arg)

    def load(self, caller, filename) -> None:
        try:
            with open(filename, "r") as f:
                N = yaml.load(f, Loader=yaml.Loader)
                # N = nx.read_yaml(filename)
                self.init(N, filename)
        except Exception as ex:
            debug("file open failed. no infection to current graph.")
            raise NetworkError("load file failed:{}".format(filename), ex)

    def save(self, caller, args) -> None:
        try:
            if len(args) == 0:
                if self._filename is None:
                    debug("filename not specified")
                else:
                    with open(self.filename, "w") as f:
                        yaml.dump(self.N, f)
                    # nx.write_yaml(self.N, self.filename)
                    # self.dirty = False
            else:
                filename = args[0]
                filename = os.getcwd() + "\\" + "{}".format(filename)
                with open(filename, "w") as f:
                    yaml.dump(self.N, f)
                # nx.write_yaml(self.N, filename)
                self._filename = filename
        except Exception as ex:
            debug("file save failed. no infection to current graph.")
            raise NetworkError("save file failed:({})".format(args), ex)

    def getcwd(self, caller):
        try:
            path = os.getcwd()
            return path
        except Exception as ex:
            raise NetworkError("couldn't getcwd", ex)

    def chdir(self, caller, dir):
        try:
            os.chdir(dir)
        except Exception as ex:
            raise NetworkError("couldn't chdir:{}".format(dir), ex)

    def modelcheck(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("modelcheck", caller, args)
                if ret is not None:
                    return ret
            # assumes sufficient arguments
            opt_spec = args[0]
            spec = opt_spec.value
            if isinstance(spec, NetworkSymbol):
                spec = spec.value
            opt_cmd = args[1]
            if opt_cmd.name in ("construct", "con"):
                self.construct_spec_network(caller, spec)
                rtn = NetworkReturnValue(spec, True, "successed")
                return rtn
            elif opt_cmd.name in ("reach", "reachability"):
                opt_act = args[2]
                if opt_act.name in ("segment", "seg"):
                    if opt_act.has_assignee:
                        seg = opt_act.value
                        if isinstance(seg, NetworkSymbol):
                            spec = seg.value
                    else:
                        seg = Numberset(1)
                        seg.infinitize()
                    opt_dep = args[3]
                    if opt_dep.has_assignee:
                        depth = opt_dep.value
                        if isinstance(depth, NetworkSymbol):
                            depth = depth.value
                    else:
                        depth = Numberset(1)
                        depth.infinitize()
                    rtn = self.analyze_reach_segments(caller, spec, seg, depth)
                    if rtn is None:
                        return NetworkReturnValue(None, False, "Unknown error")
                    rtn = NetworkReturnValue(rtn, True, "success")
                    return rtn
                elif opt_act.name in ("loopback", "lb"):
                    _from = args[3].value
                    _to = args[4].value
                    opt_from = args[3]
                    if opt_from.name in ("from"):
                        if opt_from.has_assignee:
                            _from = opt_from.value
                            if isinstance(_from, NetworkSymbol):
                                _from = _from.value
                        else:
                            _from = Numberset(1)
                            _from.infinitize()
                    opt_to = args[4]
                    if opt_to.has_assignee:
                        _to = opt_to.value
                        if isinstance(_to, NetworkSymbol):
                            _to = _to.value
                    else:
                        _to = Numberset(1)
                        _to.infinitize()
                    rtn = self.analyze_reach_loopback(caller, spec, _from, _to)
                    if rtn is None:
                        return NetworkReturnValue(None, False, "Unknown error")
                    rtn = NetworkReturnValue(rtn, True, "success")
                    return rtn
                else:
                    rtn = NetworkReturnValue(None, False, "Unknown sub-command '{}' for command 'reach'".format(opt_act.name))
                    return rtn
            else:
                rtn = NetworkReturnValue(None, False, "Unknown command:{}".format(opt_cmd.name))
                return rtn

        except Exception as ex:
            debug("modelcheck error with args:{}".format(args))
            raise NetworkError("modelcheck error:{}".format(args), ex)

    def set_node_attrib(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("import_xlsx", caller, args)
            if ret is not None:
                return ret
        attrib = args[2].value
        opt = CommandOption("-data", False, has_assignee=True)
        N = self.collect_nodes(caller, (args[0], args[1], opt))
        for n in N:
            for a in attrib:
                self.N.nodes[n][a.l] = a.r
                debug(n, "->", a.l, "=", a.r)
        return N

    def import_xlsx(self, caller, args):
        # "graph.xlsx", (Nodes, start_row, id_col, attr1, attr2,..),
        #               (Edges, start_row, src_id_col, dst_id_col, attr1, attr2,..)
        try:
            if len(args) > 0:
                ret = self.help_check("import_xlsx", caller, args)
                if ret is not None:
                    return ret
            filename = args[0]
            wb = openpyxl.load_workbook(filename, data_only=True)
            # load nodes
            nodes = []
            dataset = args[1]
            sheet = dataset[0]
            ws = wb[sheet]
            id_col = dataset[1]
            start_row = dataset[2]
            start_col = 3
            i = start_row
            while True:
                id_cell = ws.cell(row=i, column=id_col)
                if id_cell.value is None:
                    break
                attrs = {}
                for j in range(start_col, len(dataset)):
                    j = dataset[j]
                    attrs[ws.cell(row=1, column=j).value] = ws.cell(row=i, column=j).value
                nodes.append((id_cell.value, attrs))
                i += 1
            # load nodes
            edges = []
            dataset = args[2]
            sheet = dataset[0]
            ws = wb[sheet]
            src_id_col = dataset[1]
            dst_id_col = dataset[2]
            start_row = dataset[3]
            start_col = 4
            i = start_row
            while True:
                src_id_cell = ws.cell(row=i, column=src_id_col)
                dst_id_cell = ws.cell(row=i, column=dst_id_col)
                if src_id_cell.value is None or dst_id_cell.value is None:
                    break
                attrs = {}
                for j in range(start_col, len(dataset)):
                    j = dataset[j]
                    attrs[ws.cell(row=1, column=j).value] = ws.cell(row=i, column=j).value
                edges.append((src_id_cell.value, dst_id_cell.value, attrs))
                i += 1
            # import nodes
            for n in nodes:
                self.newnode(caller, (n[0], n[1], True))
            # import edges
            for e in edges:
                self.newedge(caller, (e[0], e[1], e[2], True))
            return nodes, edges
        except Exception as ex:
            raise NetworkError("excel file import error:{}".format(args), ex)

    def load_xlsx(self, caller, args):
        # "something.xlsx", (sheet1, start_row, attr1, attr2,..),
        #                   (sheet2, start_row, attr1, attr2,..),
        #                    ....
        try:
            if len(args) > 0:
                ret = self.help_check("load_xlsx", caller, args)
                if ret is not None:
                    return ret
            filename = args[0].value
            if filename[0] == "\"" and filename[len(filename)-1] == "\"":
                filename = filename[1:len(filename)-1]
            wb = openpyxl.load_workbook(filename, data_only=True)
            # load nodes
            tables = []
            for t in args[1:]:
                t = t.value
                sheet = t[0]
                cols = t[1:]
                ws = wb[sheet]
                table = []
                i = 1
                row = []
                for c in cols:
                    row.append(ws.cell(row=i, column=c).value)
                row = tuple(row)
                table.append(row)
                i = 2
                lasted = False
                while not lasted:
                    row = []
                    for c in cols:
                        cell = ws.cell(row=i, column=c)
                        if cell.value is None:
                            lasted = True
                            break
                        row.append(cell.value)
                    if not lasted:
                        row = tuple(row)
                        table.append(row)
                    i = i + 1
                tables.append(table)
            return tables
        except Exception as ex:
            debug(ex)
            raise ex

    def read_text_file(self, caller, file):
        with open(file, "r") as f:
            t = f.read()
            return t

    def save_text_file(self, caller, file, t):
        with open(file, "w") as f:
            f.write(t)
            return t

    def sort(self, caller, args):
        try:
            if len(args) > 0:
                ret = self.help_check("sort", caller, args)
                if ret is not None:
                    return ret
            sortee = args[0]
            sorter = DefaultSorter(sortee)
            comparator = Comparator()
            result = sorter.sort(comparator)
            return result
        except Exception as ex:
            raise NetworkError("sort error:{}".format(args), ex)

    def intersect(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("intersect", caller, args)
            if ret is not None:
                return ret
        U = args[0]
        V = args[1]
        I = []
        for u in U:
            if u in V:
                I.append(u)
        return I

    def union(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("union", caller, args)
            if ret is not None:
                return ret
        pass

    def pull(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("pull", caller, args)
            if ret is not None:
                return ret
        C = args[0]
        if len(C) > 0:
            c = C[0]
            C.pop(0)
            return c
        else:
            return None

    def push(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("push", caller, args)
            if ret is not None:
                return ret
        C = args[0]
        c = args[1]
        C.append(c)
        return c, len(C)-1

    def pop(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("pop", caller, args)
            if ret is not None:
                return ret
        C = args[0]
        c = C[len(C)-1]
        return c, len(C)-1

    def extend(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("extend", caller, args)
            if ret is not None:
                return ret
        A = args[0]
        B = args[1]
        A.extend(B)
        return A

    def concat(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("concat", caller, args)
            if ret is not None:
                return ret
        ret = ""
        for arg in args:
            ret = "{}{}".format(ret, arg)
        return ret

    def contains(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("concat", caller, args)
            if ret is not None:
                return ret
        C = args[0].value
        e = args[1].value
        return e in C

    def keys(self, caller, args):
        if len(args) > 0:
            ret = self.help_check("keys", caller, args)
            if ret is not None:
                return ret
        C = args[0]
        K = []
        for k in C.keys():
            K.append(k)
        return tuple(K)
